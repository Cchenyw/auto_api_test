import openpyxl


class ExcelHandler:
    def __init__(self, file):
        self.file = file

    def open_excel(self, sheet_name):
        """open excel file, get sheet"""
        wb = openpyxl.load_workbook(self.file)
        # get sheet from sheet name
        sheet = wb[sheet_name]
        return sheet

    def get_columns(self, sheet_name):
        """get header(table header)"""
        wb = self.open_excel(sheet_name)
        columns = []
        # 遍历第一行
        for i in wb[1]:
            # 将遍历出来的表头字段加入到列表
            columns.append(i.value)
        return columns

    def read_excel(self, sheet_name):
        """get all, skip second row"""
        sheet = self.open_excel(sheet_name)
        rows = list(sheet.rows)[2:]
        data = []
        for row in rows:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
                data_dict = dict(zip(self.get_columns(sheet_name), row_data))
            data.append(data_dict)
        return data

    @staticmethod
    def write_excel(file, sheet_name, row, column, data):
        """write to excel"""
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheet_name]
        sheet.cell(row, column).value = data
        wb.save(file)
        wb.close()


if __name__ == "__main__":
    excel = ExcelHandler('/Users/chenyw/Desktop/小陈的工作做不完/learning/api_test_excel_template.xlsx')
    print(excel.read_excel('Sheet1'))
